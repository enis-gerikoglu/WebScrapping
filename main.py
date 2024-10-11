import os

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


def excele_yazdir(eskisehirtoplam_bisiklet_ort, eskisehirtoplam_surus_ort, eskisehirtoplam_ortsure, sakaryatoplam_bisiklet_ort,
                  sakaryatoplam_surus_ort, sakaryatoplam_ortsure, izmirtoplam_bisiklet_ort, izmirtoplam_surus_ort, izmirtoplam_ortsure,
                  sbs_without_percentage_list, summed_yss_list, summed_sgs_list, toplam_uye_sayisi, yeni_katilim,
                  eklenen_kullanici_degisim, summed_destek_talebi,  yss_total_lists ,sgs_total_lists,izmir_bisikletort_lists,eskisehir_bisikletort_lists,sakarya_bisikletort_lists,izmir_surus_lists,eskisehir_surus_lists,sakarya_surus_lists,izmir_ort_sure_lists,eskisehir_ort_sure_lists,sakarya_ort_sure_lists,sbs_izmir,sbs_eskisehir,sbs_sakarya,yeni_katilan_lists,destek_sayisi_lists):
    # Verileri düzenliyoruz
    data = {
        'Bilgi': [
            'Acilis ve Kullanim Tarifesi(sirasiyla)',
            'Sahadaki Bisiklet Sayisi',
            'Yapilan Surus Sayisi',
            'Yapilan Surus Sayisi onceki haftaya gore degisim',
            'Suruslerde gecen sure',
            'Gunluk bisiklet basi ortalama suresi',
            'Gunluk bisiklet basi yapilan surus sayisi',
            'Bir surusun ortalama Suresi'
        ],
        'DeğerEskişehir': [
            0,
            sbs_without_percentage_list[2],
            summed_yss_list[6],
            0,
            summed_sgs_list[6],
            eskisehirtoplam_bisiklet_ort,
            eskisehirtoplam_surus_ort,
            eskisehirtoplam_ortsure,
        ],
        'Yeni Sütun 1 Eskişehir': [
            0,
            sbs_eskisehir[0],
            yss_total_lists[0][6],
            0,
            sgs_total_lists[0][6],
            eskisehir_bisikletort_lists[0],
            eskisehir_surus_lists[0],
            eskisehir_ort_sure_lists[0],
        ],
        'Yeni Sütun 2 Eskişehir': [
            0,
            sbs_eskisehir[1],
            yss_total_lists[1][6],
            0,
            sgs_total_lists[1][6],
            eskisehir_bisikletort_lists[1],
            eskisehir_surus_lists[1],
            eskisehir_ort_sure_lists[1],
        ],
        'DeğerSakarya': [
            0,
            sbs_without_percentage_list[3],
            summed_yss_list[7],
            0,
            summed_sgs_list[7],
            sakaryatoplam_bisiklet_ort,
            sakaryatoplam_surus_ort,
            sakaryatoplam_ortsure,
        ],
        'Yeni Sütun 1 Sakarya': [
            0,
            sbs_sakarya[0],
            yss_total_lists[0][7],
            0,
            sgs_total_lists[0][7],
            sakarya_bisikletort_lists[0],
            sakarya_surus_lists[0],
            sakarya_ort_sure_lists[0]
        ],
        'Yeni Sütun 2 Sakarya': [
            0,
            sbs_sakarya[1],
            yss_total_lists[1][7],
            0,
            sgs_total_lists[1][7],
            sakarya_bisikletort_lists[1],
            sakarya_surus_lists[1],
            sakarya_ort_sure_lists[1]
        ],
        'Değerİzmir': [
            0,
            sbs_without_percentage_list[1],
            summed_yss_list[5],
            0,
            summed_sgs_list[5],
            izmirtoplam_bisiklet_ort,
            izmirtoplam_surus_ort,
            izmirtoplam_ortsure,
        ],
        'Yeni Sütun 1 İzmir': [
            0,
            sbs_izmir[0],
            yss_total_lists[0][5],
            0,
            sgs_total_lists[0][5],
            izmir_bisikletort_lists[0],
            izmir_surus_lists[0],
            izmir_ort_sure_lists[0]
        ],
        'Yeni Sütun 2 İzmir': [
            0,
            sbs_izmir[1],
            yss_total_lists[1][5],
            0,
            sgs_total_lists[1][5],
            izmir_bisikletort_lists[1],
            izmir_surus_lists[1],
            izmir_ort_sure_lists[1]
        ]
    }

    # DataFrame oluşturma
    df = pd.DataFrame({
        'Bilgi': data['Bilgi'],
        'DeğerEskişehir': data['DeğerEskişehir'],
        'Ilk Hafta': data['Yeni Sütun 1 Eskişehir'],
        'Son Hafta': data['Yeni Sütun 2 Eskişehir']
    })

    # Excel'e yazdırma
    df.to_excel('rapor.xlsx', index=False)

    # Excel dosyasını yükleyip aktif sayfayı seçme
    wb = load_workbook('rapor.xlsx')
    ws = wb.active

    fill_sari = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Genel İstatistik başlığı
    ws.merge_cells('A1')
    ws['A1'] = 'Genel İstatistik'
    ws['A1'].alignment = Alignment(horizontal='center')  # Ortalı yapıyoruz
    ws['A1'].fill = fill_sari  # Sarı renge boyama
    ws['C1'] = 'İlk Hafta'
    ws['D1'] = 'Son Hafta'
    ws['B1'] = 'Toplam'

    # Genel İstatistik değerleri
    genel_istatistikler = [
        ('Toplam Üye Sayısı', toplam_uye_sayisi, toplam_uye_sayisi, toplam_uye_sayisi),
        # 3. ve 4. sütunlar için eklemeler
        ('Yeni Katılan Üye', yeni_katilim, yeni_katilan_lists[0], yeni_katilan_lists[1]),  # İki ayrı değer girilmiş
        ('Önceki Haftaya Göre Eklenen Kullanıcı Sayısı Değişimi', eklenen_kullanici_degisim, eklenen_kullanici_degisim,
         eklenen_kullanici_degisim),  # Eksik sütunlar doldurulmuş
        ('Destek Talebi', summed_destek_talebi, destek_sayisi_lists[0], destek_sayisi_lists[1])  # Eksik sütunlar doldurulmuş
    ]

    # Verileri hücrelere yazdırma
    for idx, (bilgi, deger, ilk_hafta_deger, son_hafta_deger) in enumerate(genel_istatistikler, start=2):
        ws.cell(row=idx, column=1, value=bilgi)  # Bilgi sütunu
        ws.cell(row=idx, column=2, value=deger)  # Değer sütunu
        ws.cell(row=idx, column=3, value=ilk_hafta_deger)  # 3. sütun (İlk Hafta)
        ws.cell(row=idx, column=4, value=son_hafta_deger)  # 4. sütun (Son Hafta)



    # Eskişehir başlığı
    start_row_eskisehir = 6  # Genel İstatistikten sonra başlıyor
    ws.merge_cells(start_row=start_row_eskisehir, start_column=1, end_row=start_row_eskisehir, end_column=1)
    ws.cell(row=start_row_eskisehir, column=1, value='Eskişehir').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_eskisehir, column=1).fill = fill_sari

    # Eskişehir verilerini ekleme
    for r in range(len(df)):
        ws.cell(row=start_row_eskisehir + r + 1, column=1, value=df['Bilgi'][r])
        ws.cell(row=start_row_eskisehir + r + 1, column=2, value=df['DeğerEskişehir'][r])
        ws.cell(row=start_row_eskisehir + r + 1, column=3, value=df['Ilk Hafta'][r])
        ws.cell(row=start_row_eskisehir + r + 1, column=4, value=df['Son Hafta'][r])

    # Sakarya başlığı
    start_row_sakarya = start_row_eskisehir + len(df) + 1
    ws.merge_cells(start_row=start_row_sakarya, start_column=1, end_row=start_row_sakarya, end_column=1)
    ws.cell(row=start_row_sakarya, column=1, value='Sakarya').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_sakarya, column=1).fill = fill_sari

    # Sakarya verilerini ekleme
    for r in range(len(data['Bilgi'])):
        ws.cell(row=start_row_sakarya + r + 1, column=1, value=data['Bilgi'][r])
        ws.cell(row=start_row_sakarya + r + 1, column=2, value=data['DeğerSakarya'][r])
        ws.cell(row=start_row_sakarya + r + 1, column=3, value=data['Yeni Sütun 1 Sakarya'][r])
        ws.cell(row=start_row_sakarya + r + 1, column=4, value=data['Yeni Sütun 2 Sakarya'][r])

    # İzmir başlığı
    start_row_izmir = start_row_sakarya + len(data['Bilgi']) + 1
    ws.merge_cells(start_row=start_row_izmir, start_column=1, end_row=start_row_izmir, end_column=1)
    ws.cell(row=start_row_izmir, column=1, value='İzmir').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_izmir, column=1).fill = fill_sari

    # İzmir verilerini ekleme
    for r in range(len(data['Bilgi'])):
        ws.cell(row=start_row_izmir + r + 1, column=1, value=data['Bilgi'][r])
        ws.cell(row=start_row_izmir + r + 1, column=2, value=data['Değerİzmir'][r])
        ws.cell(row=start_row_izmir + r + 1, column=3, value=data['Yeni Sütun 1 İzmir'][r])
        ws.cell(row=start_row_izmir + r + 1, column=4, value=data['Yeni Sütun 2 İzmir'][r])

    # Sütun genişliklerini ayarlama
    ws.column_dimensions['A'].width = 50  # Bilgi sütunu genişliği
    ws.column_dimensions['B'].width = 20  # Değer sütunu genişliği
    ws.column_dimensions['C'].width = 20  # Yeni Sütun 1 genişliği
    ws.column_dimensions['D'].width = 20  # Yeni Sütun 2 genişliği

    # Excel dosyasını kaydetme
    masaustu_dizini = os.path.expanduser("~/Desktop")

    # Rapor dosyasının tam yolu
    rapor_yolu = os.path.join(masaustu_dizini, 'rapor.xlsx')

    # Raporu kaydet
    wb.save(rapor_yolu)


    print("Veriler başarıyla Masa Üstünde Excel'e aktarıldı.")



def excele_yazdir1(eskisehirtoplam_bisiklet_ort, eskisehirtoplam_surus_ort, eskisehirtoplam_ortsure, sakaryatoplam_bisiklet_ort,
                  sakaryatoplam_surus_ort, sakaryatoplam_ortsure, izmirtoplam_bisiklet_ort, izmirtoplam_surus_ort, izmirtoplam_ortsure,
                  sbs_without_percentage_list, summed_yss_list, summed_sgs_list, toplam_uye_sayisi, yeni_katilim,
                  eklenen_kullanici_degisim, summed_destek_talebi,  yss_total_lists ,sgs_total_lists,izmir_bisikletort_lists,eskisehir_bisikletort_lists,sakarya_bisikletort_lists,izmir_surus_lists,eskisehir_surus_lists,sakarya_surus_lists,izmir_ort_sure_lists,eskisehir_ort_sure_lists,sakarya_ort_sure_lists,sbs_izmir,sbs_eskisehir,sbs_sakarya,yeni_katilan_lists,destek_sayisi_lists):
    # Verileri düzenliyoruz
    data = {
        'Bilgi': [
            'Acilis ve Kullanim Tarifesi(sirasiyla)',
            'Sahadaki Bisiklet Sayisi',
            'Yapilan Surus Sayisi',
            'Yapilan Surus Sayisi onceki haftaya gore degisim',
            'Suruslerde gecen sure',
            'Gunluk bisiklet basi ortalama suresi',
            'Gunluk bisiklet basi yapilan surus sayisi',
            'Bir surusun ortalama Suresi'
        ],
        'DeğerEskişehir': [
            0,
            sbs_without_percentage_list[2],
            summed_yss_list[6],
            0,
            summed_sgs_list[6],
            eskisehirtoplam_bisiklet_ort,
            eskisehirtoplam_surus_ort,
            eskisehirtoplam_ortsure,
        ],
        'Yeni Sütun 1 Eskişehir': [
            0,
            sbs_eskisehir[0],
            yss_total_lists[0][6],
            0,
            sgs_total_lists[0][6],
            eskisehir_bisikletort_lists[0],
            eskisehir_surus_lists[0],
            eskisehir_ort_sure_lists[0],
        ],

        'DeğerSakarya': [
            0,
            sbs_without_percentage_list[3],
            summed_yss_list[7],
            0,
            summed_sgs_list[7],
            sakaryatoplam_bisiklet_ort,
            sakaryatoplam_surus_ort,
            sakaryatoplam_ortsure,
        ],
        'Yeni Sütun 1 Sakarya': [
            0,
            sbs_sakarya[0],
            yss_total_lists[0][7],
            0,
            sgs_total_lists[0][7],
            sakarya_bisikletort_lists[0],
            sakarya_surus_lists[0],
            sakarya_ort_sure_lists[0]
        ],

        'Değerİzmir': [
            0,
            sbs_without_percentage_list[1],
            summed_yss_list[5],
            0,
            summed_sgs_list[5],
            izmirtoplam_bisiklet_ort,
            izmirtoplam_surus_ort,
            izmirtoplam_ortsure,
        ],
        'Yeni Sütun 1 İzmir': [
            0,
            sbs_izmir[0],
            yss_total_lists[0][5],
            0,
            sgs_total_lists[0][5],
            izmir_bisikletort_lists[0],
            izmir_surus_lists[0],
            izmir_ort_sure_lists[0]
        ],

    }

    # DataFrame oluşturma
    df = pd.DataFrame({
        'Bilgi': data['Bilgi'],
        'DeğerEskişehir': data['DeğerEskişehir'],
        'Ilk Hafta': data['Yeni Sütun 1 Eskişehir'],

    })

    # Excel'e yazdırma
    df.to_excel('rapor.xlsx', index=False)

    # Excel dosyasını yükleyip aktif sayfayı seçme
    wb = load_workbook('rapor.xlsx')
    ws = wb.active

    fill_sari = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Genel İstatistik başlığı
    ws.merge_cells('A1')
    ws['A1'] = 'Genel İstatistik'
    ws['A1'].alignment = Alignment(horizontal='center')  # Ortalı yapıyoruz
    ws['A1'].fill = fill_sari  # Sarı renge boyama
    ws['C1'] = 'İlk Hafta'

    ws['B1'] = 'Toplam'

    # Genel İstatistik değerleri
    genel_istatistikler = [
        ('Toplam Üye Sayısı', toplam_uye_sayisi, toplam_uye_sayisi),
        # 3. ve 4. sütunlar için eklemeler
        ('Yeni Katılan Üye', yeni_katilim, yeni_katilan_lists[0], ),  # İki ayrı değer girilmiş
        ('Önceki Haftaya Göre Eklenen Kullanıcı Sayısı Değişimi', eklenen_kullanici_degisim, eklenen_kullanici_degisim),  # Eksik sütunlar doldurulmuş
        ('Destek Talebi', summed_destek_talebi, destek_sayisi_lists[0], )  # Eksik sütunlar doldurulmuş
    ]

    # Verileri hücrelere yazdırma
    for idx, (bilgi, deger, ilk_hafta_deger, ) in enumerate(genel_istatistikler, start=2):
        ws.cell(row=idx, column=1, value=bilgi)  # Bilgi sütunu
        ws.cell(row=idx, column=2, value=deger)  # Değer sütunu
        ws.cell(row=idx, column=3, value=ilk_hafta_deger)  # 3. sütun (İlk Hafta)
         # 4. sütun (Son Hafta)



    # Eskişehir başlığı
    start_row_eskisehir = 6  # Genel İstatistikten sonra başlıyor
    ws.merge_cells(start_row=start_row_eskisehir, start_column=1, end_row=start_row_eskisehir, end_column=1)
    ws.cell(row=start_row_eskisehir, column=1, value='Eskişehir').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_eskisehir, column=1).fill = fill_sari

    # Eskişehir verilerini ekleme
    for r in range(len(df)):
        ws.cell(row=start_row_eskisehir + r + 1, column=1, value=df['Bilgi'][r])
        ws.cell(row=start_row_eskisehir + r + 1, column=2, value=df['DeğerEskişehir'][r])
        ws.cell(row=start_row_eskisehir + r + 1, column=3, value=df['Ilk Hafta'][r])


    # Sakarya başlığı
    start_row_sakarya = start_row_eskisehir + len(df) + 1
    ws.merge_cells(start_row=start_row_sakarya, start_column=1, end_row=start_row_sakarya, end_column=1)
    ws.cell(row=start_row_sakarya, column=1, value='Sakarya').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_sakarya, column=1).fill = fill_sari

    # Sakarya verilerini ekleme
    for r in range(len(data['Bilgi'])):
        ws.cell(row=start_row_sakarya + r + 1, column=1, value=data['Bilgi'][r])
        ws.cell(row=start_row_sakarya + r + 1, column=2, value=data['DeğerSakarya'][r])
        ws.cell(row=start_row_sakarya + r + 1, column=3, value=data['Yeni Sütun 1 Sakarya'][r])


    # İzmir başlığı
    start_row_izmir = start_row_sakarya + len(data['Bilgi']) + 1
    ws.merge_cells(start_row=start_row_izmir, start_column=1, end_row=start_row_izmir, end_column=1)
    ws.cell(row=start_row_izmir, column=1, value='İzmir').alignment = Alignment(horizontal='center')
    ws.cell(row=start_row_izmir, column=1).fill = fill_sari

    # İzmir verilerini ekleme
    for r in range(len(data['Bilgi'])):
        ws.cell(row=start_row_izmir + r + 1, column=1, value=data['Bilgi'][r])
        ws.cell(row=start_row_izmir + r + 1, column=2, value=data['Değerİzmir'][r])
        ws.cell(row=start_row_izmir + r + 1, column=3, value=data['Yeni Sütun 1 İzmir'][r])


    # Sütun genişliklerini ayarlama
    ws.column_dimensions['A'].width = 50  # Bilgi sütunu genişliği
    ws.column_dimensions['B'].width = 20  # Değer sütunu genişliği
    ws.column_dimensions['C'].width = 20  # Yeni Sütun 1 genişliği
    ws.column_dimensions['D'].width = 20  # Yeni Sütun 2 genişliği

    # Excel dosyasını kaydetme
    masaustu_dizini = os.path.expanduser("~/Desktop")

    # Rapor dosyasının tam yolu
    rapor_yolu = os.path.join(masaustu_dizini, 'rapor.xlsx')

    # Raporu kaydet
    wb.save(rapor_yolu)

    print("Veriler başarıyla Masa Üstünde Excel'e aktarıldı.")
def main():
      sbs_izmir = []
      sbs_sakarya = []
      sbs_eskisehir = []
      izmir_bisikletort_lists = []
      eskisehir_bisikletort_lists = []
      sakarya_bisikletort_lists = []
      izmir_surus_lists = []
      eskisehir_surus_lists = []
      sakarya_surus_lists = []
      izmir_ort_sure_lists = []
      eskisehir_ort_sure_lists = []
      sakarya_ort_sure_lists = []
      sbs_without_percentage_lists = []
      yeni_katilan_lists = []
      yss_total_lists = []
      sgs_total_minutes_lists = []
      destek_sayisi_lists = []
      bisiklet_ort_lists = []
      surus_ort_lists = []
      ortsure_lists = []
      urls = []
      num_links = int(input("Kaç tane link girmek istersiniz? "))
      urls = [input(f"Link {i + 1}: ") for i in range(num_links)]





      for url in urls:

        urlss = url
        response = requests.get(urlss)
        soup = BeautifulSoup(response.content, 'html.parser')
        #butun display-6 spanlarini bulma
        menu_elements = soup.find_all('span', class_='display-6')
        #destek talebi sayisi bulma
        destek_sayisi = int(menu_elements[3].text.strip())
        destek_sayisi_lists.append(destek_sayisi)
        #yeni katilan uye sayisi bulma
        yeni_katilan_tr = int(menu_elements[0].text.strip())
        yeni_katilan_yabanci = int(menu_elements[1].text.strip())
        yeni_katilan_toplam = yeni_katilan_yabanci + yeni_katilan_tr
        yeni_katilan_lists.append(yeni_katilan_toplam)





        #butun table elementlerini bulma
        table_elements = soup.find_all('td')

        #yssleri bulma ve toplama
        td_yss_locations = [4,13,31,40,49,58,67,76]
        yss_total_list = []

        for yssidx in td_yss_locations:
            yssvalue_int = int(table_elements[yssidx].text.strip())
            yss_total_list.append(yssvalue_int)

        yss_total_sum = sum(yss_total_list)

        #surus dakikalarini bulma ve toplama
        def calculate_total_minutes(time_str):
            timeparts = time_str.split(':')

            if len(timeparts) == 3:
                hours, minutes, seconds = timeparts
            elif len(timeparts) == 2:
                hours = '0'
                minutes, seconds = timeparts
            elif len(timeparts) == 1:
                hours = '0'
                minutes = '0'
                seconds = timeparts[0]

            total_minutes = int(hours) * 60 + int(minutes)

            return total_minutes

        td_sgg_locations = [6, 15, 33, 42, 51, 60, 69, 78]
        sgs_total_minutes_list = []

        for idx in td_sgg_locations:
            time_str = str(table_elements[idx].text.strip())
            total_minutes = calculate_total_minutes(time_str)
            sgs_total_minutes_list.append(total_minutes)

        sgs_total_sum = sum(sgs_total_minutes_list)


        #sbs bulma
        td_sbs_locations = [11,56,65,74]
        sbspercentage_list = []
        sbs_without_percentage_list = []

        for sbspercidx in td_sbs_locations:
            sbsvalue_str = str(table_elements[sbspercidx].text.strip())
            sbspercentage_list.append(sbsvalue_str)

        for sbs in sbspercentage_list:
            beforpercentage = sbs.split('%')[0].strip()
            beforpercentage_int = int(beforpercentage)
            sbs_without_percentage_list.append(beforpercentage_int)


        sbs_izmir.append(sbs_without_percentage_list[1])
        sbs_eskisehir.append(sbs_without_percentage_list[2])
        sbs_sakarya.append(sbs_without_percentage_list[3])

        izmir_bisiklet_ort = sgs_total_minutes_list[5] / sbs_without_percentage_list[1]
        eskisehir_bisiklet_ort = sgs_total_minutes_list[6] / sbs_without_percentage_list[2]
        sakarya_bisiklet_ort = sgs_total_minutes_list[7] / sbs_without_percentage_list[3]

        izmir_bisikletort_lists.append(izmir_bisiklet_ort)
        eskisehir_bisikletort_lists.append(eskisehir_bisiklet_ort)
        sakarya_bisikletort_lists.append(sakarya_bisiklet_ort)

        # gunluk bisiklet basi yapilan surus sayisi
        izmir_surus_ort = yss_total_list[5] / sbs_without_percentage_list[1]
        eskisehir_surus_ort = yss_total_list[6] / sbs_without_percentage_list[2]
        sakarya_surus_ort = yss_total_list[7] / sbs_without_percentage_list[3]

        izmir_surus_lists.append(izmir_surus_ort)
        eskisehir_surus_lists.append(eskisehir_surus_ort)
        sakarya_surus_lists.append(sakarya_surus_ort)

        # bir surusun ortalama suresini bulma
        izmir_ortsure = sgs_total_minutes_list[5] / yss_total_list[5]
        eskisehir_ortsure = sgs_total_minutes_list[6] / yss_total_list[6]
        sakarya_ortsure = sgs_total_minutes_list[7] / yss_total_list[7]

        izmir_ort_sure_lists.append(izmir_ortsure)
        eskisehir_ort_sure_lists.append(eskisehir_ortsure)
        sakarya_ort_sure_lists.append(sakarya_ortsure)
        #bisiklet basina ortalama dk bulma


        #bisiklet_ort_lists.append(izmir_bisiklet_ort)
        #bisiklet_ort_lists.append(eskisehir_bisiklet_ort)
        #bisiklet_ort_lists.append(sakarya_bisiklet_ort)

        #surus_ort_lists.append(izmir_surus_ort)
        #surus_ort_lists.append(eskisehir_surus_ort)
        #surus_ort_lists.append(sakarya_surus_ort)

        #ortsure_lists.append(izmir_ortsure)
        #ortsure_lists.append(eskisehir_ortsure)
        #ortsure_lists.append(sakarya_ortsure)

        sbs_without_percentage_lists.append(sbs_without_percentage_list)
        yss_total_lists.append(yss_total_list)
        sgs_total_minutes_lists.append(sgs_total_minutes_list)





      summed_sbs_list = [sum(values) for values in zip(*sbs_without_percentage_lists)]

      summed_yenikatilan = sum(yeni_katilan_lists)
      summed_destek_talebi = sum(destek_sayisi_lists)

      summed_yss_list = [sum(values) for values in zip(*yss_total_lists)]

      summed_sgs_list = [sum(values) for values in zip(*sgs_total_minutes_lists)]




      izmirtoplam_bisiklet_ort = summed_sgs_list[5] / summed_sbs_list[1]
      eskisehirtoplam_bisiklet_ort = summed_sgs_list[6] / summed_sbs_list[2]
      sakaryatoplam_bisiklet_ort = summed_sgs_list[7] / summed_sbs_list[3]

      # gunluk bisiklet basi yapilan surus sayisi
      izmirtoplam_surus_ort = summed_yss_list[5] / summed_sbs_list[1]
      eskisehirtoplam_surus_ort = summed_yss_list[6] / summed_sbs_list[2]
      sakaryatoplam_surus_ort = summed_yss_list[7] / summed_sbs_list[3]


      # bir surusun ortalama suresini bulma
      izmirtoplam_ortsure = summed_sgs_list[5] / summed_yss_list[5]
      eskisehirtoplam_ortsure = summed_sgs_list[6] / summed_yss_list[6]
      sakaryatoplam_ortsure = summed_sgs_list[7] / summed_yss_list[7]




      if len(urls) == 2:

          excele_yazdir(eskisehirtoplam_bisiklet_ort, eskisehirtoplam_surus_ort, eskisehirtoplam_ortsure, sakaryatoplam_bisiklet_ort,
                        sakaryatoplam_surus_ort, sakaryatoplam_ortsure, izmirtoplam_bisiklet_ort, izmirtoplam_surus_ort, izmirtoplam_ortsure,
                        sbs_without_percentage_list, summed_yss_list, summed_sgs_list, 0, summed_yenikatilan, 0,
                        summed_destek_talebi, yss_total_lists, sgs_total_minutes_lists,izmir_bisikletort_lists,eskisehir_bisikletort_lists,sakarya_bisikletort_lists,izmir_surus_lists,eskisehir_surus_lists,sakarya_surus_lists,izmir_ort_sure_lists,eskisehir_ort_sure_lists,sakarya_ort_sure_lists,sbs_izmir,sbs_eskisehir,sbs_sakarya,yeni_katilan_lists,destek_sayisi_lists)
      else:
          excele_yazdir1(eskisehirtoplam_bisiklet_ort, eskisehirtoplam_surus_ort, eskisehirtoplam_ortsure, sakaryatoplam_bisiklet_ort,
                        sakaryatoplam_surus_ort, sakaryatoplam_ortsure, izmirtoplam_bisiklet_ort, izmirtoplam_surus_ort, izmirtoplam_ortsure,
                        sbs_without_percentage_list, summed_yss_list, summed_sgs_list, 0, summed_yenikatilan, 0,
                        summed_destek_talebi, yss_total_lists, sgs_total_minutes_lists,izmir_bisikletort_lists,eskisehir_bisikletort_lists,sakarya_bisikletort_lists,izmir_surus_lists,eskisehir_surus_lists,sakarya_surus_lists,izmir_ort_sure_lists,eskisehir_ort_sure_lists,sakarya_ort_sure_lists,sbs_izmir,sbs_eskisehir,sbs_sakarya,yeni_katilan_lists,destek_sayisi_lists)









if __name__ == "__main__":
    main()